# src/tool/make_paper_tables.py
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

from src.utils.paths import project_root

# ============================================================
# Config
# ============================================================
SQI_ORDER = ["bSQI", "iSQI", "kSQI", "sSQI", "pSQI", "fSQI", "basSQI"]
GROUP_ORDER = ["Pairs", "Triplets", "Quadruplets", "Quintuplets", "Sextuplets", "All SQI"]

SETTING_ORDER = ["trainA_testB", "trainB_testA"] 
SETTING_TO_HEADER = {
    "trainA_testB": ("Training Set-a", "Test Set-b"),
    "trainB_testA": ("Training Set-b", "Test Set-a"),
}

# ============================================================
# Helpers
# ============================================================
def _read_csv(p: Path) -> pd.DataFrame:
    if not p.exists():
        raise FileNotFoundError(f"Missing file: {p}")
    df = pd.read_csv(p)
    if len(df) == 0:
        raise ValueError(f"Empty CSV: {p}")
    return df

def _fmt_float(x: Any, nd: int = 3) -> str:
    try:
        if pd.isna(x):
            return ""
        return f"{float(x):.{nd}f}"
    except Exception:
        return ""

def _ensure_outdir(out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

def _sort_sqi_cols(cols: list[str]) -> list[str]:
    order = {k: i for i, k in enumerate(SQI_ORDER)}
    return sorted(cols, key=lambda c: order.get(c, 999))

# ============================================================
# Table 5
# ============================================================
def make_table5_paper_wide(df_long: pd.DataFrame, nd: int = 3) -> pd.DataFrame:
    """
    Input long (one row per SQI):
      SQI, Ac_train, Se_train, Sp_train, Ac_test, Se_test, Sp_test
    Output wide like paper:
                bSQI  iSQI  ... basSQI
      Train Set-a‡ Ac  ...
                  Se
                  Sp
      Test  Set-b‡ Ac
                  Se
                  Sp
    """
    df = df_long.copy()
    if "SQI" not in df.columns:
        raise ValueError("Table5 needs column 'SQI'")

    # enforce SQI order
    df["SQI"] = df["SQI"].astype(str)
    df["_o"] = df["SQI"].map({k: i for i, k in enumerate(SQI_ORDER)}).fillna(999).astype(int)
    df = df.sort_values("_o").drop(columns=["_o"])

    # build row blocks
    row_spec = [
        ("Train\nSet-a\u2021", "Ac", "Ac_train"),
        ("", "Se", "Se_train"),
        ("", "Sp", "Sp_train"),
        ("Test\nSet-b\u2021", "Ac", "Ac_test"),
        ("", "Se", "Se_test"),
        ("", "Sp", "Sp_test"),
    ]

    # create a wide table with 2 index columns: block label + metric label
    out = pd.DataFrame(index=pd.MultiIndex.from_tuples([(a, b) for a, b, _ in row_spec],
                                                      names=["", ""]),
                       columns=_sort_sqi_cols(df["SQI"].tolist()))

    for block_label, metric_label, col in row_spec:
        for _, r in df.iterrows():
            sqi = r["SQI"]
            if sqi not in out.columns:
                continue
            out.loc[(block_label, metric_label), sqi] = _fmt_float(r.get(col, None), nd)

    # make it look closer to paper when exported: keep MultiIndex rows
    return out

# ============================================================
# Table 6
# ============================================================
def make_table6_paper(df: pd.DataFrame, nd: int = 3) -> pd.DataFrame:
    d = df.copy()
    for need in ["Group", "Selected_SQI", "Ac_train", "Ac_test"]:
        if need not in d.columns:
            raise ValueError(f"Table6 needs column '{need}'")

    d["Group"] = d["Group"].astype(str)
    d["_o"] = d["Group"].map({k: i for i, k in enumerate(GROUP_ORDER)}).fillna(999).astype(int)
    d = d.sort_values("_o").drop(columns=["_o"]).reset_index(drop=True)

    out = pd.DataFrame({
        "Selected SQI": d["Selected_SQI"].astype(str).to_numpy(),
        "Ac training\nSet-a\u2021": d["Ac_train"].apply(lambda x: _fmt_float(x, nd)).to_numpy(),
        "Ac test\nSet-b\u2021": d["Ac_test"].apply(lambda x: _fmt_float(x, nd)).to_numpy(),
    })
    out.index = d["Group"].to_numpy()
    out.index.name = ""
    return out

# ============================================================
# Table 7
# ============================================================
def _normalize_setting(s: str) -> str:
    """
    Map your internal names to two canonical settings used in Table7.
    Adjust this if your CSV uses different names.
    """
    s = str(s)
    # common guesses:
    if s in ("trainA_testB", "A2B", "A_to_B", "SetA_to_SetB"):
        return "trainA_testB"
    if s in ("trainB_testA", "B2A", "B_to_A", "SetB_to_SetA"):
        return "trainB_testA"
    return s

def make_table7_paper(mlp_df: pd.DataFrame, svm_df: Optional[pd.DataFrame], nd: int = 3) -> pd.DataFrame:
    """
    Expect each df to contain rows for the 4 metrics, plus balanced versions (‡).
    Minimal required cols (per row):
      Setting, balanced (0/1 or False/True), Ac_test, Se_test, Sp_test, (optional) PNet
    We'll lay out columns:
      Training Set-a: MLP SVM | Test Set-b: MLP SVM | Training Set-b: MLP SVM | Test Set-a: MLP SVM
    Rows:
      Ac / Se / Sp / PNet, then Ac‡ / Se‡ / Sp‡ / PNet‡
    """
    def prep(df: pd.DataFrame, model_name: str) -> pd.DataFrame:
        d = df.copy()
        if "Setting" not in d.columns:
            raise ValueError(f"Table7({model_name}) needs column 'Setting'")

        # if absent, assume all unbalanced
        if "balanced" not in d.columns:
            d["balanced"] = 0

        d["Setting_raw"] = d["Setting"].astype(str)
        d["Setting"] = d["Setting"].apply(_normalize_setting)

        # ---- NEW: if normalization didn't produce trainA_testB/trainB_testA, map uniques deterministically ----
        known = {"trainA_testB", "trainB_testA"}
        uniq = list(d["Setting"].astype(str).unique())

        if not any(u in known for u in uniq):
            raw_uniq = list(d["Setting_raw"].astype(str).unique())

            if len(raw_uniq) == 1:
                # only one setting in file -> treat as trainA_testB
                d["Setting"] = "trainA_testB"
            elif len(raw_uniq) >= 2:
                # first -> A2B, second -> B2A
                mapping = {raw_uniq[0]: "trainA_testB", raw_uniq[1]: "trainB_testA"}
                d["Setting"] = d["Setting_raw"].map(mapping).fillna("trainA_testB")

        return d

    mlp = prep(mlp_df, "MLP")
    svm = prep(svm_df, "SVM") if svm_df is not None else None

    # Build empty output
    col_tuples = [
        ("Training Set-a", "MLP"), ("Training Set-a", "SVM"),
        ("Test Set-b", "MLP"), ("Test Set-b", "SVM"),
        ("Training Set-b", "MLP"), ("Training Set-b", "SVM"),
        ("Test Set-a", "MLP"), ("Test Set-a", "SVM"),
    ]
    cols = pd.MultiIndex.from_tuples(col_tuples, names=["", ""])
    row_index = ["Ac", "Se", "Sp", "PNet", "Ac\u2021", "Se\u2021", "Sp\u2021", "PNet\u2021"]
    out = pd.DataFrame("", index=row_index, columns=cols)

    def fill_one(model_df: pd.DataFrame, model_label: str) -> None:
        for balanced_flag, suffix in [(0, ""), (1, "\u2021")]:
            sub = model_df[model_df["balanced"].astype(int) == balanced_flag]

            for setting in SETTING_ORDER:
                sub_s = sub[sub["Setting"] == setting]
                if len(sub_s) == 0:
                    continue
                # if multiple rows, take first (you can make it stricter if you want)
                r = sub_s.iloc[0]

                # map to the 2 column groups
                train_hdr, test_hdr = SETTING_TO_HEADER.get(setting, (setting, setting))

                # choose train metrics columns if exist, else fall back to test (best effort)
                ac_tr = r.get("Ac_train", r.get("Ac_test", None))
                se_tr = r.get("Se_train", r.get("Se_test", None))
                sp_tr = r.get("Sp_train", r.get("Sp_test", None))

                ac_te = r.get("Ac_test", None)
                se_te = r.get("Se_test", None)
                sp_te = r.get("Sp_test", None)

                # Optional PNet column name variants
                pnet = None
                for key in ["PNet", "pnet", "PNet_test"]:
                    if key in r.index:
                        pnet = r.get(key, None)
                        break

                out.loc[f"Ac{suffix}", (train_hdr, model_label)] = _fmt_float(ac_tr, nd)
                out.loc[f"Se{suffix}", (train_hdr, model_label)] = _fmt_float(se_tr, nd)
                out.loc[f"Sp{suffix}", (train_hdr, model_label)] = _fmt_float(sp_tr, nd)

                out.loc[f"Ac{suffix}", (test_hdr, model_label)] = _fmt_float(ac_te, nd)
                out.loc[f"Se{suffix}", (test_hdr, model_label)] = _fmt_float(se_te, nd)
                out.loc[f"Sp{suffix}", (test_hdr, model_label)] = _fmt_float(sp_te, nd)

                if pnet is not None:
                    out.loc[f"PNet{suffix}", (test_hdr, model_label)] = _fmt_float(pnet, nd)

    fill_one(mlp, "MLP")
    if svm is not None:
        fill_one(svm, "SVM")

    out.index.name = ""
    return out

# ============================================================
# Underline best cells in Excel (simple rule)
# ============================================================
def _underline_max_in_rows(ws, start_row: int, start_col: int, nrows: int, ncols: int, row_names: set[str]) -> None:
    """
    A simple underline rule: for selected row labels (e.g., {"Ac", "Ac‡", "PNet", "PNet‡"}),
    underline the max numeric value in that row across the data range.
    """
    underline_font = Font(underline="single")

    for r in range(start_row, start_row + nrows):
        row_label = ws.cell(row=r, column=start_col).value
        if row_label not in row_names:
            continue

        vals = []
        for c in range(start_col + 1, start_col + 1 + ncols):
            v = ws.cell(row=r, column=c).value
            try:
                fv = float(v)
                vals.append((fv, c))
            except Exception:
                pass
        if not vals:
            continue
        maxv = max(v for v, _ in vals)
        for v, c in vals:
            if abs(v - maxv) < 1e-12:
                ws.cell(row=r, column=c).font = underline_font

# ============================================================
# Main
# ============================================================
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Make paper-aligned tables (wide) from SVM/MLP outputs")
    p.add_argument("--seed", type=int, default=0)
    p.add_argument("--artifacts_dir", type=str, default="artifacts")
    p.add_argument("--out_dir", type=str, default="artifacts/paper_tables")
    p.add_argument("--nd", type=int, default=3)
    return p.parse_args()

def main() -> None:
    args = parse_args()
    seed = int(args.seed)
    nd = int(args.nd)

    root = project_root()
    artifacts = root / args.artifacts_dir
    out_dir = root / args.out_dir
    _ensure_outdir(out_dir)

    svm_dir = artifacts / "models" / "svm"
    mlp_dir = artifacts / "models" / "lm_mlp" / "tables"

    svm_t5 = svm_dir / f"table5_12lead_single_sqi_seed{seed}.csv"
    svm_t6 = svm_dir / f"table6_12lead_combo_sqi_seed{seed}.csv"
    svm_t7 = svm_dir / f"table7_svm_selected5_seed{seed}.csv"  # 建议你产出这个

    mlp_t5 = mlp_dir / f"table5_mlp_12lead_single_sqi_seed{seed}.csv"
    mlp_t6 = mlp_dir / f"table6_mlp_12lead_combo_sqi_seed{seed}.csv"
    mlp_t7 = mlp_dir / f"table7_mlp_selected5_seed{seed}.csv"

    outputs: dict[str, pd.DataFrame] = {}

    # ---- Table 5 ----
    if svm_t5.exists():
        outputs["Table5_SVM"] = make_table5_paper_wide(_read_csv(svm_t5), nd=nd)
    if mlp_t5.exists():
        outputs["Table5_MLP"] = make_table5_paper_wide(_read_csv(mlp_t5), nd=nd)

    # ---- Table 6 ----
    if svm_t6.exists():
        outputs["Table6_SVM"] = make_table6_paper(_read_csv(svm_t6), nd=nd)
    if mlp_t6.exists():
        outputs["Table6_MLP"] = make_table6_paper(_read_csv(mlp_t6), nd=nd)

    # ---- Table 7 (merge) ----
    if mlp_t7.exists():
        mlp_df = _read_csv(mlp_t7)
        svm_df = _read_csv(svm_t7) if svm_t7.exists() else None
        outputs["Table7_MLP+SVM"] = make_table7_paper(mlp_df, svm_df, nd=nd)

    if not outputs:
        raise FileNotFoundError("No input tables found. Run your svm/mlp pipelines that export tables first.")

    # save CSVs (note: MultiIndex will be written with extra header rows, OK)
    for name, df in outputs.items():
        p = out_dir / f"{name}.csv"
        df.to_csv(p)
        print(f"[saved] {p}")

    # save Excel with better look + optional underline
    xlsx_path = out_dir / "paper_tables_like_paper.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        for name, df in outputs.items():
            sheet = name[:31]
            df.to_excel(w, sheet_name=sheet)

        wb = w.book
        # Optional underline: Table5 underline best in Ac rows; Table6 underline best accuracies; Table7 underline best in Ac/PNet rows.
        for sheet in wb.worksheets:
            title = sheet.title
            if title.startswith("Table5_"):
                # find data range roughly: MultiIndex rows -> first col is blank, second col metric label
                # We'll underline max in the two "Ac" rows only (best-effort)
                _underline_max_in_rows(
                    sheet,
                    start_row=2, start_col=1,
                    nrows=6, ncols=7,
                    row_names={"Ac"}
                )
            if title.startswith("Table6_"):
                # underline max in both accuracy columns across groups (best-effort)
                # Here first column is index (Group), then Selected SQI, then 2 acc columns
                # We'll just underline max within each acc column not implemented (keep simple)
                pass
            if title.startswith("Table7_"):
                _underline_max_in_rows(
                    sheet,
                    start_row=2, start_col=1,
                    nrows=8, ncols=8,
                    row_names={"Ac", "Ac\u2021", "PNet", "PNet\u2021"}
                )

    print(f"[saved] {xlsx_path}")

if __name__ == "__main__":
    main()